"""Device management endpoints — the heart of the portal."""

import csv
import io
import json
import secrets
from datetime import datetime, timedelta, timezone
from typing import Annotated, Optional
from fastapi import APIRouter, Depends, HTTPException, Query, Request, Response, status
from fastapi.responses import StreamingResponse
from io import BytesIO
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.api.deps import (
    get_current_user,
    get_db,
    require_admin_or_operator,
)
from app.core.crypto import encrypt_app_key, decrypt_app_key
from app.db.models.audit import AuditLog
from app.db.models.device import Device
from app.db.models.device_type import DeviceType
from app.db.models.user import User
from app.schemas.device import (
    DeviceCreate,
    DeviceCreatedResponse,
    DeviceListResponse,
    DeviceOut,
    DeviceTypeOut,
    DeviceUpdate,
)
from app.services.chirpstack_client import cs_client, ChirpStackError
from app.services.influxdb_client import influx_client
from app.services.sticker_generator import generate_sticker_pdf

router = APIRouter()


# ============ Device Types catalog ============

@router.get("/types", response_model=list[DeviceTypeOut])
async def list_device_types(
    user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> list[DeviceTypeOut]:
    """List all available device types (catalog of supported sensors)."""
    result = await db.execute(select(DeviceType).order_by(DeviceType.name))
    types = result.scalars().all()
    return [
        DeviceTypeOut(
            id=str(t.id),
            name=t.name,
            manufacturer=t.manufacturer,
            model=t.model,
            region=t.region,
            description=t.description,
        )
        for t in types
    ]


# ============ Device CRUD ============

@router.get("", response_model=DeviceListResponse)
async def list_devices(
    user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    search: Optional[str] = Query(None),
    status_filter: Optional[str] = Query(None, alias="status"),
) -> DeviceListResponse:
    """List devices for the authenticated user's customer account."""
    # Base query — strictly scoped to this customer
    base = (
        select(Device)
        .options(selectinload(Device.device_type))
        .where(Device.customer_account_id == user.customer_account_id)
    )

    if search:
        search_term = f"%{search.lower()}%"
        base = base.where(
            (func.lower(Device.name).like(search_term))
            | (func.lower(Device.dev_eui).like(search_term))
        )

    if status_filter:
        base = base.where(Device.status == status_filter)

    # Count total
    count_q = select(func.count()).select_from(base.subquery())
    total = (await db.execute(count_q)).scalar() or 0

    # Paginate
    devices_q = base.order_by(Device.created_at.desc()).offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(devices_q)
    devices = result.scalars().all()

    return DeviceListResponse(
        items=[_device_to_out(d) for d in devices],
        total=total,
        page=page,
        page_size=page_size,
    )


@router.post(
    "",
    response_model=DeviceCreatedResponse,
    status_code=status.HTTP_201_CREATED,
)
async def create_device(
    payload: DeviceCreate,
    request: Request,
    db: Annotated[AsyncSession, Depends(get_db)],
    user: Annotated[User, Depends(require_admin_or_operator)],
) -> DeviceCreatedResponse:
    """
    Create a new device. THE critical endpoint.

    Flow:
        1. Validate inputs and device type
        2. Generate a fresh AppKey
        3. Create the device in ChirpStack (creates + sets keys)
        4. Save device record in portal DB
        5. If portal DB save fails, roll back ChirpStack
        6. Return the device + plaintext AppKey ONCE (for sticker printing)
    """
    # Look up device type
    device_type = await db.get(DeviceType, payload.device_type_id)
    if not device_type:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unknown device type",
        )

    # Check DevEUI not already in use across the system
    existing = await db.execute(
        select(Device).where(Device.dev_eui == payload.dev_eui)
    )
    if existing.scalar_one_or_none():
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="A device with this DevEUI already exists",
        )

    account = user.customer_account
    if not account.chirpstack_application_id:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Customer account is missing ChirpStack application — contact support",
        )

    # Generate a fresh 128-bit AppKey (32 hex chars)
    app_key = secrets.token_hex(16)

    # Step 1: Create in ChirpStack first. If it fails, no portal cleanup needed.
    try:
        cs_client.create_device(
            dev_eui=payload.dev_eui,
            join_eui=payload.join_eui,
            app_key=app_key,
            application_id=account.chirpstack_application_id,
            device_profile_id=device_type.chirpstack_profile_id,
            name=payload.name,
        )
    except ChirpStackError as e:
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail=f"Failed to provision in ChirpStack: {e}",
        )

    # Step 2: Save in portal DB. If this fails, roll back ChirpStack.
    try:
        device = Device(
            customer_account_id=account.id,
            device_type_id=device_type.id,
            dev_eui=payload.dev_eui,
            join_eui=payload.join_eui,
            app_key_encrypted=encrypt_app_key(app_key),
            name=payload.name,
            location_name=payload.location_name,
            location_lat=payload.location_lat,
            location_lon=payload.location_lon,
            status="pending",
            created_by=user.id,
        )
        db.add(device)

        audit = AuditLog(
            user_id=user.id,
            customer_account_id=account.id,
            action="device.create",
            target_type="device",
            target_id=payload.dev_eui,
            ip=request.client.host if request.client else None,
            user_agent=request.headers.get("user-agent", "")[:500],
            details={
                "name": payload.name,
                "device_type": device_type.name,
                "location": payload.location_name,
            },
        )
        db.add(audit)

        await db.commit()
        await db.refresh(device, ["device_type"])

    except Exception as e:
        # Roll back: delete from ChirpStack
        try:
            cs_client.delete_device(payload.dev_eui)
        except ChirpStackError:
            pass  # log but don't propagate; primary error is more important
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Database error: {e}",
        )

    # Step 3: Return with plaintext AppKey for the sticker (ONE TIME)
    return DeviceCreatedResponse(
        **_device_to_out(device).model_dump(),
        app_key=app_key,
        sticker_url=f"/api/v1/devices/{device.dev_eui}/sticker?app_key={app_key}",
    )


@router.get("/{dev_eui}", response_model=DeviceOut)
async def get_device(
    dev_eui: str,
    user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> DeviceOut:
    """Get details for a single device. Enforces customer ownership."""
    device = await _load_device_for_user(db, dev_eui, user)
    return _device_to_out(device)


@router.patch("/{dev_eui}", response_model=DeviceOut)
async def update_device(
    dev_eui: str,
    payload: DeviceUpdate,
    request: Request,
    db: Annotated[AsyncSession, Depends(get_db)],
    user: Annotated[User, Depends(require_admin_or_operator)],
) -> DeviceOut:
    """Update device metadata (name, location)."""
    device = await _load_device_for_user(db, dev_eui, user)

    changes = {}
    for field in ("name", "location_name", "location_lat", "location_lon", "status"):
        value = getattr(payload, field, None)
        if value is not None:
            changes[field] = value
            setattr(device, field, value)

    audit = AuditLog(
        user_id=user.id,
        customer_account_id=user.customer_account_id,
        action="device.update",
        target_type="device",
        target_id=dev_eui,
        ip=request.client.host if request.client else None,
        details={"changes": changes},
    )
    db.add(audit)
    await db.commit()
    await db.refresh(device, ["device_type"])

    return _device_to_out(device)


@router.delete("/{dev_eui}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_device(
    dev_eui: str,
    request: Request,
    db: Annotated[AsyncSession, Depends(get_db)],
    user: Annotated[User, Depends(require_admin_or_operator)],
) -> None:
    """Delete a device from both ChirpStack and the portal."""
    device = await _load_device_for_user(db, dev_eui, user)

    # Delete from ChirpStack first
    try:
        cs_client.delete_device(dev_eui)
    except ChirpStackError as e:
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail=f"Failed to delete from ChirpStack: {e}",
        )

    audit = AuditLog(
        user_id=user.id,
        customer_account_id=user.customer_account_id,
        action="device.delete",
        target_type="device",
        target_id=dev_eui,
        ip=request.client.host if request.client else None,
        details={"name": device.name},
    )
    db.add(audit)

    await db.delete(device)
    await db.commit()


# ============ Sticker generation ============

@router.get("/{dev_eui}/sticker")
async def get_device_sticker(
    dev_eui: str,
    user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    app_key: Optional[str] = Query(None, description="Plaintext AppKey from creation response"),
) -> Response:
    """
    Generate a printable sticker PDF for a device.

    Two modes:
    1. Immediately after creation: pass app_key as query param (already shown to user)
    2. Later request: app_key is decrypted from DB (admin-only, audited)
    """
    device = await _load_device_for_user(db, dev_eui, user)

    # Determine which AppKey to encode
    if app_key:
        # Trust the caller — this is the freshly-created sticker request
        sticker_app_key = app_key
    else:
        # Decrypt stored AppKey. Only admins should hit this path.
        if user.role != "admin":
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Only admins may reprint stickers after creation",
            )
        sticker_app_key = decrypt_app_key(device.app_key_encrypted)

        # Audit log this access
        audit = AuditLog(
            user_id=user.id,
            customer_account_id=user.customer_account_id,
            action="device.sticker_reprint",
            target_type="device",
            target_id=dev_eui,
        )
        db.add(audit)
        await db.commit()

    pdf_bytes = generate_sticker_pdf(
        dev_eui=device.dev_eui,
        app_key=sticker_app_key,
        join_eui=device.join_eui,
        device_name=device.name,
        device_type=f"{device.device_type.manufacturer or ''} {device.device_type.model or ''}".strip()
                    or device.device_type.name,
    )

    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'inline; filename="sticker-{dev_eui}.pdf"',
        },
    )


# ============ Device measurements (time series) ============

@router.get("/{dev_eui}/measurements")
async def get_device_measurements(
    dev_eui: str,
    user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    from_time: Optional[str] = Query(None, description="ISO datetime, e.g. 2024-01-01T00:00:00"),
    to_time: Optional[str] = Query(None, description="ISO datetime, e.g. 2024-01-31T23:59:59"),
    hours: int = Query(24, ge=1, le=8760),
    field: Optional[str] = Query(None),
) -> dict:
    """Get historical readings for a device from InfluxDB.

    Pass from_time + to_time for a specific range, or just hours for a rolling window.
    """
    device = await _load_device_for_user(db, dev_eui, user)
    ft, tt = _resolve_time_range(from_time, to_time, hours)

    measurements = influx_client.get_device_measurements(
        dev_eui=device.dev_eui,
        from_time=ft,
        to_time=tt,
        field=field,
    )

    return {
        "dev_eui": device.dev_eui,
        "from": ft.isoformat(),
        "to": tt.isoformat(),
        "count": len(measurements),
        "measurements": measurements,
    }


@router.get("/{dev_eui}/export")
async def export_device_data(
    dev_eui: str,
    user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    fmt: str = Query("csv", description="Export format: csv | json | xlsx"),
    from_time: Optional[str] = Query(None),
    to_time: Optional[str] = Query(None),
    hours: int = Query(24, ge=1, le=8760),
    field: Optional[str] = Query(None),
) -> Response:
    """Export device measurements as CSV, JSON, or XLSX."""
    if fmt not in ("csv", "json", "xlsx"):
        raise HTTPException(status_code=400, detail="fmt must be csv, json, or xlsx")

    device = await _load_device_for_user(db, dev_eui, user)
    ft, tt = _resolve_time_range(from_time, to_time, hours)

    rows = influx_client.get_device_measurements(
        dev_eui=device.dev_eui,
        from_time=ft,
        to_time=tt,
        field=field,
        limit=50000,
    )

    # Normalise timestamps to ISO strings for export
    normalised = [
        {
            "timestamp": r["timestamp"].isoformat() if hasattr(r["timestamp"], "isoformat") else str(r["timestamp"]),
            "field": r["field"],
            "value": r["value"],
        }
        for r in rows
    ]

    filename = f"{dev_eui}_{ft.strftime('%Y%m%d')}_{tt.strftime('%Y%m%d')}"

    if fmt == "json":
        content = json.dumps(
            {
                "dev_eui": device.dev_eui,
                "device_name": device.name,
                "from": ft.isoformat(),
                "to": tt.isoformat(),
                "count": len(normalised),
                "measurements": normalised,
            },
            indent=2,
        ).encode()
        return Response(
            content=content,
            media_type="application/json",
            headers={"Content-Disposition": f'attachment; filename="{filename}.json"'},
        )

    if fmt == "csv":
        out = io.StringIO()
        writer = csv.writer(out)
        writer.writerow(["timestamp", "field", "value"])
        for r in normalised:
            writer.writerow([r["timestamp"], r["field"], r["value"]])
        return Response(
            content=out.getvalue().encode(),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{filename}.csv"'},
        )

    # xlsx
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Measurements"

    # Header row
    headers = ["Timestamp", "Field", "Value"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for i, r in enumerate(normalised, 2):
        ws.cell(row=i, column=1, value=r["timestamp"])
        ws.cell(row=i, column=2, value=r["field"])
        ws.cell(row=i, column=3, value=r["value"])

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 14

    # Metadata sheet
    meta = wb.create_sheet("Info")
    meta.append(["Device EUI", device.dev_eui])
    meta.append(["Device name", device.name])
    meta.append(["From", ft.isoformat()])
    meta.append(["To", tt.isoformat()])
    meta.append(["Total rows", len(normalised)])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
    )


@router.get("/{dev_eui}/latest")
async def get_device_latest(
    dev_eui: str,
    user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict:
    """Get the latest reading for a device."""
    device = await _load_device_for_user(db, dev_eui, user)
    latest = influx_client.get_latest_reading(device.dev_eui)
    return {
        "dev_eui": device.dev_eui,
        "latest": latest,
    }


# ============ Helpers ============

async def _load_device_for_user(
    db: AsyncSession,
    dev_eui: str,
    user: User,
) -> Device:
    """Load a device and enforce that it belongs to the user's account."""
    result = await db.execute(
        select(Device)
        .options(selectinload(Device.device_type))
        .where(Device.dev_eui == dev_eui.lower())
        .where(Device.customer_account_id == user.customer_account_id)
    )
    device = result.scalar_one_or_none()
    if not device:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Device not found",
        )
    return device


def _resolve_time_range(
    from_time: Optional[str],
    to_time: Optional[str],
    hours: int,
) -> tuple[datetime, datetime]:
    """Return (from, to) datetime pair. from_time/to_time take priority over hours."""
    now = datetime.now(timezone.utc).replace(tzinfo=None)
    if from_time:
        try:
            ft = datetime.fromisoformat(from_time.replace("Z", ""))
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid from_time format — use ISO 8601")
    else:
        ft = now - timedelta(hours=hours)

    if to_time:
        try:
            tt = datetime.fromisoformat(to_time.replace("Z", ""))
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid to_time format — use ISO 8601")
    else:
        tt = now

    if ft >= tt:
        raise HTTPException(status_code=400, detail="from_time must be before to_time")

    return ft, tt


def _device_to_out(device: Device) -> DeviceOut:
    """Convert ORM Device to DeviceOut response."""
    return DeviceOut(
        id=str(device.id),
        dev_eui=device.dev_eui,
        join_eui=device.join_eui,
        name=device.name,
        location_name=device.location_name,
        location_lat=float(device.location_lat) if device.location_lat is not None else None,
        location_lon=float(device.location_lon) if device.location_lon is not None else None,
        status=device.status,
        last_seen_at=device.last_seen_at,
        created_at=device.created_at,
        device_type=DeviceTypeOut(
            id=str(device.device_type.id),
            name=device.device_type.name,
            manufacturer=device.device_type.manufacturer,
            model=device.device_type.model,
            region=device.device_type.region,
            description=device.device_type.description,
        ),
    )
